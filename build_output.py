"""Merge Stage 1 + Stage 2 results and build the output workbook.

Stage 2's answer wins for any row it processed; Stage 1's stands otherwise.
Adds resolved_by_stage (1 or 2). Builds Detail / Summary / Not Captured sheets
per Claude.md.
"""

import json
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

from gemini_common import load_templates

STAGE1_CHECKPOINT = "stage1_checkpoint.jsonl"
STAGE2_CHECKPOINT = "stage2_checkpoint.jsonl"
OUTPUT_PATH = "Output Excel Sheets/Actual_To_Atomic_Mapping.xlsx"


def load_jsonl(path):
    with open(path, encoding="utf-8") as f:
        return [json.loads(l) for l in f]


def merge():
    stage1_rows = {r["row_number"]: r for r in load_jsonl(STAGE1_CHECKPOINT)}
    stage2_rows = {r["row_number"]: r for r in load_jsonl(STAGE2_CHECKPOINT)}

    merged = []
    for row_number, s1 in stage1_rows.items():
        if row_number in stage2_rows:
            r = dict(stage2_rows[row_number])
            r["resolved_by_stage"] = 2
        else:
            r = dict(s1)
            r["resolved_by_stage"] = 1
        merged.append(r)

    merged.sort(key=lambda r: r["row_number"])
    return merged


def tags_to_str(tags):
    return "; ".join(f"{t['tag']}={t['value']} ({t['confidence']})" for t in tags)


def build_detail_sheet(wb, merged):
    ws = wb.active
    ws.title = "Detail"
    headers = ["row_number", "comment_text", "count", "matched_template_id",
               "match_confidence", "extracted_tags", "notes", "resolved_by_stage"]
    ws.append(headers)
    for r in merged:
        ws.append([
            r["row_number"], r["comment_text"], r["count"], r["matched_template_id"],
            r["match_confidence"], tags_to_str(r["extracted_tags"]), r.get("notes", ""),
            r["resolved_by_stage"],
        ])
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 2)
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["F"].width = 50


def build_summary_sheet(wb, merged, templates):
    ws = wb.create_sheet("Summary")
    ws.append(["Template ID", "Unique Comments Covered", "Total Occurrences Covered", "Distinct Tag-Value Count"])

    by_template = defaultdict(list)
    for r in merged:
        by_template[r["matched_template_id"]].append(r)

    template_ids = [t["template_id"] for t in templates] + ["No match"]
    for tid in template_ids:
        rows = by_template.get(tid, [])
        unique_comments = len(rows)
        total_occurrences = sum(r["count"] for r in rows)
        distinct_tag_values = len({
            (t["tag"], t["value"]) for r in rows for t in r["extracted_tags"]
        })
        ws.append([tid, unique_comments, total_occurrences, distinct_tag_values])

    for col, width in zip("ABCD", [16, 24, 26, 24]):
        ws.column_dimensions[col].width = width


def build_not_captured_sheet(wb, merged):
    ws = wb.create_sheet("Not Captured")
    headers = ["row_number", "comment_text", "count", "match_confidence", "resolved_by_stage", "notes"]
    ws.append(headers)
    not_captured = [r for r in merged if r["matched_template_id"] == "No match"]
    not_captured.sort(key=lambda r: r["count"], reverse=True)
    for r in not_captured:
        ws.append([r["row_number"], r["comment_text"], r["count"], r["match_confidence"],
                   r["resolved_by_stage"], r.get("notes", "")])
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["F"].width = 50
    return len(not_captured)


def main():
    merged = merge()
    templates = load_templates()

    wb = openpyxl.Workbook()
    build_detail_sheet(wb, merged)
    build_summary_sheet(wb, merged, templates)
    not_captured_count = build_not_captured_sheet(wb, merged)

    wb.save(OUTPUT_PATH)

    total_unique = len(merged)
    total_occurrences = sum(r["count"] for r in merged)
    matched = [r for r in merged if r["matched_template_id"] != "No match"]
    matched_unique = len(matched)
    matched_occurrences = sum(r["count"] for r in matched)
    resolved_by_2 = sum(1 for r in merged if r["resolved_by_stage"] == 2)

    stage1_rows = {r["row_number"]: r for r in load_jsonl(STAGE1_CHECKPOINT)}
    changed_by_stage2 = 0
    for r in merged:
        if r["resolved_by_stage"] == 2:
            s1 = stage1_rows[r["row_number"]]
            if s1["matched_template_id"] != r["matched_template_id"]:
                changed_by_stage2 += 1

    print(f"Total unique comments: {total_unique}")
    print(f"Total occurrence volume: {total_occurrences}")
    print(f"Matched unique comments: {matched_unique} ({matched_unique/total_unique*100:.2f}%)")
    print(f"Matched occurrence volume: {matched_occurrences} ({matched_occurrences/total_occurrences*100:.2f}%)")
    print(f"Not Captured (final No match) rows: {not_captured_count}")
    print(f"Rows resolved by Stage 2: {resolved_by_2}")
    print(f"Rows where Stage 2 changed Stage 1's matched_template_id: {changed_by_stage2}")


if __name__ == "__main__":
    main()
