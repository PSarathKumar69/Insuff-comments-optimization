"""Shared helpers for Stage 1 / Stage 2 Gemini classification scripts.

Loads the Education sheet rows, the Tag Glossary, and the T1-T68 template
list from the source workbook, and defines the structured-output schema
both stages share. See Claude.md for the full spec.
"""

import json
import os
import time
import random
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import openpyxl
from dotenv import load_dotenv
from google import genai
from google.genai import types
from google.genai import errors
from pydantic import BaseModel

SOURCE_XLSX = "Source Excel sheets/F1-ALL Strucutred Comments.xlsx"
BATCH_SIZE = 100


class ExtractedTag(BaseModel):
    tag: str
    value: str
    confidence: str  # High / Medium / Low


class RowResult(BaseModel):
    row_number: int
    comment_text: str
    count: int
    matched_template_id: str  # "T1".."T68" or "No match"
    match_confidence: str  # High / Medium / Low
    extracted_tags: List[ExtractedTag]
    notes: Optional[str] = ""


def load_client() -> genai.Client:
    load_dotenv()
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY not set in environment/.env")
    return genai.Client(api_key=api_key)


def load_rows(xlsx_path: str = SOURCE_XLSX):
    """Returns list of dicts: {row_number, comment_text, count} for every
    populated row in Education!A/B (row 2 onward)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Education"]
    rows = []
    for r in range(2, ws.max_row + 1):
        comment = ws.cell(row=r, column=1).value
        count = ws.cell(row=r, column=2).value
        if comment is None or str(comment).strip() == "":
            continue
        rows.append({
            "row_number": r,
            "comment_text": str(comment).strip(),
            "count": int(count) if count is not None else 0,
        })
    return rows


def load_templates(xlsx_path: str = SOURCE_XLSX):
    """Returns list of {template_id, template_text} for T1-T68 from
    Education!C/D."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Education"]
    templates = []
    for r in range(2, ws.max_row + 1):
        tid = ws.cell(row=r, column=3).value
        ttext = ws.cell(row=r, column=4).value
        if tid is None or str(tid).strip() == "":
            continue
        templates.append({
            "template_id": str(tid).strip(),
            "template_text": str(ttext).strip() if ttext else "",
        })
    return templates


def load_tag_glossary(xlsx_path: str = SOURCE_XLSX):
    """Returns list of {tag, glossary, examples, status} for all 36 rows."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Tag Glossary"]
    tags = []
    for r in range(2, ws.max_row + 1):
        tag = ws.cell(row=r, column=1).value
        if tag is None or str(tag).strip() == "":
            continue
        tags.append({
            "tag": str(tag).strip(),
            "glossary": ws.cell(row=r, column=2).value or "",
            "examples": ws.cell(row=r, column=3).value or "",
            "status": ws.cell(row=r, column=4).value or "",
        })
    return tags


def build_fixed_context(templates, tags) -> str:
    """Builds the fixed context block (templates + tag glossary) included
    in every batch call. Built once, reused across all calls."""
    lines = ["## Template list (T1-T68) — match by MEANING, not literal wording:\n"]
    for t in templates:
        lines.append(f"{t['template_id']}: {t['template_text']}")

    lines.append("\n## Tag Glossary — valid tags only, never invent new ones:\n")
    for t in tags:
        lines.append(f"{t['tag']}: {t['glossary']} (examples: {t['examples']})")

    return "\n".join(lines)


def make_batches(rows, batch_size: int = BATCH_SIZE):
    return [rows[i:i + batch_size] for i in range(0, len(rows), batch_size)]


def build_prompt(fixed_context: str, batch_rows) -> str:
    rows_json = json.dumps(
        [{"row_number": r["row_number"], "comment_text": r["comment_text"], "count": r["count"]}
         for r in batch_rows],
        ensure_ascii=False,
    )
    return f"""{fixed_context}

## Task

For each row below, an "Actual Comment" written by a real reviewer during an
education-verification background check:

1. Match it to the single best-fitting template (T1-T68) by MEANING/intent,
   not literal word overlap — real comments are phrased very differently from
   the clean templates. If nothing fits, use matched_template_id = "No match".
2. Set match_confidence to High, Medium, or Low based on how confident the
   semantic match is.
3. Extract every tag value present, per the Tag Glossary above. A comment can
   have 0 or several tags. Only use tags from the glossary — never invent one.
4. Add a short note only if there's something ambiguous or worth flagging.

Return the row_number, comment_text, and count exactly as given below (do not
alter them).

## Rows to classify (JSON array):
{rows_json}
"""


def call_gemini_sync(client: genai.Client, model: str, prompt: str, max_retries: int = 6):
    """Synchronous call with exponential backoff retry on 429s."""
    config = types.GenerateContentConfig(
        response_mime_type="application/json",
        response_schema=list[RowResult],
        temperature=0.1,
    )
    attempt = 0
    while True:
        try:
            response = client.models.generate_content(model=model, contents=prompt, config=config)
            return response.parsed
        except errors.ClientError as e:
            if e.code == 429 and attempt < max_retries:
                sleep_s = min(60, (2 ** attempt)) + random.uniform(0, 1)
                time.sleep(sleep_s)
                attempt += 1
                continue
            raise


async def call_gemini_async(client: genai.Client, model: str, prompt: str, max_retries: int = 6):
    """Async call with exponential backoff retry on 429s."""
    import asyncio

    config = types.GenerateContentConfig(
        response_mime_type="application/json",
        response_schema=list[RowResult],
        temperature=0.1,
    )
    attempt = 0
    while True:
        try:
            response = await client.aio.models.generate_content(model=model, contents=prompt, config=config)
            return response.parsed
        except errors.ClientError as e:
            if e.code == 429 and attempt < max_retries:
                sleep_s = min(60, (2 ** attempt)) + random.uniform(0, 1)
                await asyncio.sleep(sleep_s)
                attempt += 1
                continue
            raise
        except errors.ServerError as e:
            if attempt < max_retries:
                sleep_s = min(60, (2 ** attempt)) + random.uniform(0, 1)
                await asyncio.sleep(sleep_s)
                attempt += 1
                continue
            raise


def append_checkpoint(checkpoint_path: str, batch_index: int, results: List[RowResult]):
    """Append a completed batch's rows to the JSONL checkpoint file, then
    record the batch index as done in the sibling state file."""
    path = Path(checkpoint_path)
    with path.open("a", encoding="utf-8") as f:
        for r in results:
            row = r.model_dump()
            row["batch_index"] = batch_index
            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    state_path = path.with_suffix(".state.json")
    done = set()
    if state_path.exists():
        done = set(json.loads(state_path.read_text(encoding="utf-8")))
    done.add(batch_index)
    state_path.write_text(json.dumps(sorted(done)), encoding="utf-8")


def log_progress(message: str, log_path: str = "PROGRESS_LOG.md"):
    """Append a timestamped line to the progress log (dated entries)."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"- [{now}] {message}\n")


def load_done_batches(checkpoint_path: str) -> set:
    state_path = Path(checkpoint_path).with_suffix(".state.json")
    if not state_path.exists():
        return set()
    return set(json.loads(state_path.read_text(encoding="utf-8")))
