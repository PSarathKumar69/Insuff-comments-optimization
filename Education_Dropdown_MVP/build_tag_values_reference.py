#!/usr/bin/env python3
"""
Rebuilds Output Excel Sheets/Tag_Values_Reference.xlsx from the live
data/tag_values.json, one row per tag, sorted alphabetically. Matches the
existing workbook's schema exactly (Tag | Values currently supported | Field
type | Count | Notes / source), single sheet "Tag Values (current)", dark
navy bold header row, Arial throughout, a trailing legend row, and a "*"
suffix on the Tag name for anything new/changed/removed THIS pass.

Run from this folder: python3 build_tag_values_reference.py
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE = os.path.dirname(__file__)
tag_values = json.load(open(os.path.join(BASE, "data", "tag_values.json")))
OUT_PATH = os.path.join(BASE, "..", "Output Excel Sheets", "Tag_Values_Reference.xlsx")

# Tags touched THIS pass (2026-08-19, fifth pass - full-template tag audit):
# 4 new tags added (VISIT_DOCUMENTS for T42, PRICING_TOOL_COST/ADDITIONAL_COST
# optional cost breakdown for T8, SOURCE_FORM for T59), plus PORTAL_ACTION
# (now also wired into T17, previously only T61) and ADDRESS_TYPE (corrected -
# removed from T65, replaced with VS, after re-checking all 30 real T65 rows
# showed it was a mismatched assumption - see PROGRESS_LOG.md's 2026-08-19 entry).
# Sixth pass (2026-08-19, same day): added VISIT_REASON for T42 after the owner
# spotted the generated comment had no Reason element - see PROGRESS_LOG.md.
# Seventh pass (2026-08-21): owner flagged real duplicate values in DOCUMENTS
# ("Degree" vs "Degree certificate" - an Ops agent selecting both would produce
# a clumsy repeated comment). Removed 3 DOCUMENTS duplicates ("Degree
# certificate" -> merged into "Degree"; "Original copies of University-issued
# Final year marksheet" -> merged into "Final year marksheet"; "Letter of
# Authorization" -> merged into "Authorization form") and 1 SPECIAL_INSTRUCTIONS
# duplicate ("Duly signed by the institution", 3 real mentions -> merged into
# "Sealed and signed by the institution", 80 real mentions) - see
# PROGRESS_LOG.md's 2026-08-21 entry for the real-data evidence behind each
# merge and for 3 further clusters flagged but NOT yet merged pending owner call.
# Eighth pass (2026-08-21, same day): owner flagged a missing category (no
# semester-wise marksheet options at all - 11 real rows ask for a specific
# numbered semester or "All Semester Marksheet") and one zero-evidence value
# ("Graduation or Higher Education Documents", 0 real hits for the literal
# document phrase - the real underlying idea is a qualification-level
# descriptor, not a distinct document, already covered by existing
# Degree/marksheet entries). Added 9 semester values, removed the 1
# zero-evidence value - see PROGRESS_LOG.md's 2026-08-21 entry.
# Ninth pass (2026-08-21, same day): owner spotted "Clear and complete
# scanned/Uncut" missing from a generated comment during a live 10-comment
# demo. Grepped the dataset - this phrase appears in ~45% of ALL 8,284 real
# rows (3,724 hits), the single most common special-instruction-type phrase
# found in this project so far. Added to SPECIAL_INSTRUCTIONS.
# Tenth pass (2026-08-23, XMind revamp Step 2): ANTECEDENTS' vague "Name"
# split into "Candidate's Name" / "Candidate's Father's Name" per the XMind's
# Name Change Proof -> Mismatch branch - directly resolves the "father name
# mismatch" gap flagged earlier this session. SPECIAL_INSTRUCTIONS gained 2
# ARN-specific values, now document-conditional via new applies_to_documents
# metadata (also added to DOCUMENTS' case_types from Step 1).
# Eleventh pass (2026-08-23, XMind revamp Step 4): IDENTIFIER_TYPE gained 5
# values (Seat No./Admission No./Professional Certificate No./Verification
# No./Certificate No.) - zero real evidence, added on the XMind's authority.
# CASE_LEVEL_INFORMATION gained "Mode of Education" (0 hits for the literal
# field name, but strong evidence for its underlying values - Full time 302,
# Distance 22, Regular 12) and "Certificate Verification Link" (1 real hit).
# Twelfth pass (2026-08-23, HSC/SSC year fix): COURSE_NAME gained a year_mode
# map - HSC/SSC now render a single year instead of a from/to range (owner
# spotted an implausible "(2022-2026)" 4-year span on a generated Class-12
# comment; 409 real HSC/SSC rows checked, 112 state a single year, 0 state a
# range).
# Thirteenth pass (2026-08-23, ANTECEDENTS Mismatch-vs-Missing cleanup): owner
# reviewed a generated Info-Mismatch comment and asked which of the 12
# ANTECEDENTS values actually belong there. Removed "Bonafide/NOC" from the
# Mismatch context (it's a document, not a comparable value - restricted to
# T14/Missing only via new applies_to_templates metadata); merged "Bachelor's
# Degree" into "Education degree" (near-duplicate, only 1 real row used the
# former); renamed "School branch" to "Specialization / Branch" after finding
# 16/19 real "branch" mentions mean engineering discipline, not a campus
# location; added "Percentage / Marks Obtained" (common-sense, zero real
# evidence) and "Grade / Division" (common-sense, weak real evidence - real
# mentions exist but only as document qualifiers, not explicit mismatches).
# 2026-08-24: IDENTIFIER_TYPE and REVIEW_REASON converted dropdown_free_hybrid
# -> dropdown_multi (owner: "give multiselect options wherever required, if
# not now in future they may require so think that way as well") - both have
# a plausible real-world multi-value case (multiple missing identifiers;
# multiple simultaneous review flags) and needed zero CommentEngine.php
# changes since the generic tag-resolution loop already joins array values.
# 2026-08-24 (continued): ANTECEDENTS gained Date of Birth/Gender/Address
# (T59-scoped), CASE_LEVEL_INFORMATION gained Nationality/Board-Affiliation/
# Alternate-Contact/Photograph, FIELD_NAME gained CGPA-Percentage/Years
# Attended/Address/Identifier Number/Verification Portal Link, and a new
# INVALID_REASON tag was added for new template T72 - all part of the
# Information Reason->Scenario 2-level dropdown rebuild (owner-approved
# dropdown value expansion).
CHANGED_THIS_PASS = {"VISIT_DOCUMENTS", "PRICING_TOOL_COST", "ADDITIONAL_COST", "SOURCE_FORM", "PORTAL_ACTION", "ADDRESS_TYPE", "VISIT_REASON", "DOCUMENTS", "SPECIAL_INSTRUCTIONS", "ANTECEDENTS", "IDENTIFIER_TYPE", "CASE_LEVEL_INFORMATION", "COURSE_NAME", "REVIEW_REASON", "FIELD_NAME", "INVALID_REASON", "VERIFICATION_BLOCKER", "INCOMPLETE_DETAIL", "BLUR_DETAIL"}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tag Values (current)"

HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
BODY_FONT_BOLD = Font(name="Arial", size=10, bold=True)

headers = ["Tag", "Values currently supported", "Field type", "Count", "Notes / source"]
ws.append(headers)
for c in range(1, 6):
    cell = ws.cell(row=1, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

row_i = 2
for tag in sorted(tag_values.keys()):
    meta = tag_values[tag]
    ftype = meta.get("type", "")
    note = meta.get("note", "")
    if ftype == "system_constant":
        values_str = meta.get("value", "")
        count = 1
    elif ftype in ("free_text", "free_text_list"):
        values_str = "(unbounded — free text, not enumerated)"
        count = 0
    else:
        vals = meta.get("values", [])
        values_str = ", ".join(vals)
        count = len(vals)

    display_tag = tag + " *" if tag in CHANGED_THIS_PASS else tag
    ws.cell(row=row_i, column=1, value=display_tag).font = BODY_FONT_BOLD
    ws.cell(row=row_i, column=2, value=values_str).font = BODY_FONT
    ws.cell(row=row_i, column=3, value=ftype).font = BODY_FONT
    ws.cell(row=row_i, column=4, value=count).font = BODY_FONT
    ws.cell(row=row_i, column=5, value=note).font = BODY_FONT
    row_i += 1

row_i += 1
ws.cell(row=row_i, column=1, value="* = new, changed, or corrected as of this rebuild (2026-08-23, thirteenth pass - ANTECEDENTS Mismatch-vs-Missing cleanup)").font = Font(name="Arial", size=10, italic=True)

col_widths = {"A": 26, "B": 70, "C": 20, "D": 8, "E": 70}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w
for r in range(2, row_i):
    ws.row_dimensions[r].alignment = Alignment(wrap_text=False, vertical="top")

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}: {row_i - 3} tags ({len(CHANGED_THIS_PASS)} marked changed this pass).")
